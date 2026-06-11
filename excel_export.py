# ====================================================
# ESPORTAZIONE EXCEL E CSV ESEMPIO USO
# ====================================================
'''
from excel_export import ExcelExporter

exporter = ExcelExporter()

exporter.export_products(
    products,
    "output/codici.xlsx"
)

exporter.export_kg_products(
    products,
    "output/prodotti_kg.xlsx"
)
'''

from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from utils import Product


class ExcelExporter:
    """
    Gestione esportazioni Excel e CSV.
    """

    def __init__(self) -> None:
        pass



    def _sentence_case(
        self,
        text: str
    ) -> str:

        if not text:
            return ""

        text = " ".join(
            text.split()
        )

        return (
            text[:1].upper()
            + text[1:].lower()
        )
    # =====================================================
    # PUBLIC
    # =====================================================

    def export_products(
        self,
        products: list[Product],
        output_file: str | Path
    ) -> Path:
        """
        Esporta tutti i prodotti.
        """

        output_file = Path(output_file)

        rows = [
            {
                "Nome referenza": p.description,
                "Peso": p.weight,
                "Nome marchio": p.brand,
                "Codice": p.code,
                "Immagine": (
                    r"\\172.24.11.2\Secure\catalogue"
                    + "\\"
                    + p.code[-6:]
                    + ".jpg"
                ),
                "Logo": (
                    p.brand
                    .replace(" ", "")
                    .replace("'", "")
                    .replace("-", "")
                    .lower()
                    + ".jpg"
                ),
                "Pagina": p.page,
            }
            for p in products
        ]

        df = pd.DataFrame(
            rows,
            columns=[
                "Nome referenza",
                "Peso",
                "Nome marchio",
                "Codice",
                "Immagine",
                "Logo",
                "Pagina",
            ]
        )

        df.to_excel(
            output_file,
            index=False
        )

        self._format_excel(output_file)

        return output_file

    # =====================================================
    # KG PRODUCTS
    # =====================================================

    def export_kg_products(
        self,
        products: list[Product],
        output_file: str | Path
    ) -> Path:
        """
        Esporta solo prodotti Kg.
        """

        kg_products = [
            p for p in products
            if p.is_kg_product
        ]

        rows = [
            {
                "Nome referenza": p.description,
                "Peso": p.weight,
                "Nome marchio": p.brand,
                "Codice": p.code,
                "Immagine": (
                    r"\\172.24.11.2\Secure\catalogue"
                    + "\\"
                    + p.code[-6:]
                    + ".jpg"
                ),
                "Logo": (
                    p.brand
                    .replace(" ", "")
                    .replace("'", "")
                    .replace("-", "")
                    .lower()
                    + ".jpg"
                ),
                "Pagina": p.page,
            }
            for p in kg_products
        ]

        df = pd.DataFrame(
            rows,
            columns=[
                "Nome referenza",
                "Peso",
                "Nome marchio",
                "Codice",
                "Immagine",
                "Logo",
                "Pagina",
            ]
        )

        df.to_excel(
            output_file,
            index=False
        )

        self._format_excel(output_file)

        return output_file

    # =====================================================
    # CSV
    # =====================================================

    def export_csv(
        self,
        products: list[Product],
        output_file: str | Path
    ) -> Path:
        """
        Esporta CSV UTF-8.
        """

        rows = [
            {
                "Marchio Sezione": p.section_brand,
                "Descrizione": p.description,
                "Peso": p.weight,
                "Marchio": p.brand,
                "Codice": p.code,
                "Pagina": p.page,
            }
            for p in products
        ]

        df = pd.DataFrame(
            rows,
            columns=[
                "Marchio Sezione",
                "Descrizione",
                "Peso",
                "Marchio",
                "Codice",
                "Pagina",
            ]
        )

        df.to_csv(
            output_file,
            index=False,
            sep=";",
            encoding="utf-8-sig"
        )

        return Path(output_file)

    # =====================================================
    # BRAND REPORT
    # =====================================================

    def export_brand_report(
        self,
        brand_counter: dict,
        output_file: str | Path
    ) -> Path:
        """
        Report conteggio marchi.
        """

        rows = []

        for brand, count in sorted(
            brand_counter.items()
        ):
            rows.append(
                {
                    "Marchio": brand,
                    "Prodotti": count
                }
            )

        df = pd.DataFrame(rows)

        '''df = df[
            [
                "Section Brand",
                "Description",
                "Weight",
                "Brand",
                "Code",
                "Page"
            ]
        ]'''
        # nessun riordinamento colonne necessario

        df.to_excel(
            output_file,
            index=False
        )


        self._format_excel(output_file)

        return Path(output_file)

    # =====================================================
    # FORMATTING
    # =====================================================

    def _format_excel(
        self,
        excel_file: str | Path
    ) -> None:
        """
        Formatta workbook.
        """

        wb = load_workbook(excel_file)

        ws = wb.active

        # Header bold

        for cell in ws[1]:
            cell.font = Font(
                bold=True
            )

        # Autofit

        for column in ws.columns:

            max_length = 0

            column_letter = (
                column[0]
                .column_letter
            )

            for cell in column:

                try:
                    value = str(cell.value)

                    if len(value) > max_length:
                        max_length = len(value)

                except Exception:
                    pass

            ws.column_dimensions[
                column_letter
            ].width = max_length + 4

        wb.save(excel_file)